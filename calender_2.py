import openpyxl
from datetime import datetime, date, timedelta

def parse_dates(date_str):
    date_list = []
    for dates in date_str.split(','):
        dates = dates.strip()
        if '~' in dates:
            start, end = [datetime.strptime(x.strip(), "%Y-%m-%d").date() for x in dates.split('~')]
            while start <= end:
                date_list.append(start)
                start += timedelta(days=1)
        else:
            date_list.append(datetime.strptime(dates.strip(), "%Y-%m-%d").date())
    return date_list

# Define start_date and end_date
start_date = date(2023, 5, 1)
end_date = date(2023, 5, 31)

# Load the input data file
input_workbook = openpyxl.load_workbook("precalender.xlsx")
input_worksheet = input_workbook.active

# Load the May 2023 calendar file
output_workbook = openpyxl.load_workbook("may_2023_calendar_v2.xlsx")
output_worksheet = output_workbook["May 2023 Calendar"]

# Define the headers
spaces = ["스튜디오 블랙", "스튜디오 화이트", "무용연습실 1", "무용연습실 2", "무용연습실 3", "무용연습실 4"]
time_slots = ["오전", "오후", "야간"]

# Iterate through the input data rows (skipping the header row)
for row in input_worksheet.iter_rows(min_row=2, min_col=1, max_col=4, values_only=True):
    input_date_str, input_space, input_time, input_name = row
    input_dates = parse_dates(input_date_str)

    for input_date in input_dates:
        # Find the corresponding date column in the output worksheet
        date_column = None
        for col in range(2, (end_date - start_date).days + 3):
            if output_worksheet.cell(row=1, column=col).value == input_date.strftime("%Y-%m-%d"):
                date_column = col
                break

        # Find the corresponding space and time row in the output worksheet
        space_time_row = None
        for row in range(2, 4 * len(spaces) + 2):
            space_time_value = output_worksheet.cell(row=row, column=1).value
            if space_time_value == f"{input_space} - {input_time}":
                space_time_row = row
                break

        # Write the input name to the corresponding cell in the output worksheet
        if date_column is not None and space_time_row is not None:
            output_worksheet.cell(row=space_time_row, column=date_column).value = input_name


column_width = 10
for col in range(1, (end_date - start_date).days + 3):
    col_letter = openpyxl.utils.get_column_letter(col)
    output_worksheet.column_dimensions[col_letter].width = column_width

# Save the updated workbook to a new file
output_workbook.save("may_2023_calendar_v2_with_names.xlsx")
