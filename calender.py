import openpyxl
from openpyxl.styles import Alignment
from datetime import date, timedelta

# Create a new Excel workbook and add a worksheet
workbook = openpyxl.Workbook()
worksheet = workbook.active
worksheet.title = "May 2023 Calendar"

# Define the headers
spaces = ["스튜디오 블랙", "스튜디오 화이트", "무용연습실 1", "무용연습실 2", "무용연습실 3", "무용연습실 4"]
time_slots = ["오전", "오후", "야간"]

# Write the headers (dates)
start_date = date(2023, 5, 1)
end_date = date(2023, 5, 31)

for i in range((end_date - start_date).days + 1):
    current_date = start_date + timedelta(days=i)
    worksheet.cell(row=1, column=i + 2).value = current_date.strftime("%Y-%m-%d")
    worksheet.cell(row=1, column=i + 2).alignment = Alignment(horizontal="center")

# Write the headers (spaces and time slots)
for i, space in enumerate(spaces):
    for j, time_slot in enumerate(time_slots):
        worksheet.cell(row=i * 4 + j + 2, column=1).value = f"{space} - {time_slot}"
        worksheet.cell(row=i * 4 + j + 2, column=1).alignment = Alignment(horizontal="left")

# Set row heights
for row in range(2, 4 * len(spaces) + 2):
    if (row - 2) % 4 == 0:
        worksheet.row_dimensions[row].height = 20
    else:
        worksheet.row_dimensions[row].height = 20

# Set column widths
worksheet.column_dimensions["A"].width = 22
for col in range(2, (end_date - start_date).days + 3):
    worksheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15

# Save the workbook to a file
workbook.save("may_2023_calendar_v2.xlsx")
